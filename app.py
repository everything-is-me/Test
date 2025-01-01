from flask import Flask, render_template, request, jsonify, send_file
import os
import cv2
import mediapipe as mp
import numpy as np
import base64
from datetime import datetime
from openpyxl import Workbook
import openpyxl
from openpyxl.utils import get_column_letter


app = Flask(__name__)

# Directory for storing registered faces
REGISTER_DIR = "registered_faces"
os.makedirs(REGISTER_DIR, exist_ok=True)

mp_face_detection = mp.solutions.face_detection

def save_image_from_base64(student_name, image_data):
    """
    Save face images from Base64-encoded data using MediaPipe for face detection.
    """
    face_count = len([f for f in os.listdir(REGISTER_DIR) if f.startswith(student_name)])  # Start from the next index

    # Decode the Base64 image
    _, encoded_image = image_data.split(',', 1)
    decoded_image = base64.b64decode(encoded_image)

    # Convert the image into a NumPy array
    np_image = np.frombuffer(decoded_image, np.uint8)
    frame = cv2.imdecode(np_image, cv2.IMREAD_COLOR)

    # Process the frame with MediaPipe
    with mp_face_detection.FaceDetection(min_detection_confidence=0.5) as face_detection:
        frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        results = face_detection.process(frame_rgb)

        if results.detections:
            for detection in results.detections:
                # Extract bounding box
                bboxC = detection.location_data.relative_bounding_box
                ih, iw, _ = frame.shape
                x, y, w, h = int(bboxC.xmin * iw), int(bboxC.ymin * ih), int(bboxC.width * iw), int(bboxC.height * ih)

                # Crop the detected face
                face_crop = frame[y:y + h, x:x + w]

                # Save the cropped face
                if face_crop.size > 0:
                    face_filename = os.path.join(REGISTER_DIR, f"{student_name}_face_{face_count}.jpg")
                    cv2.imwrite(face_filename, face_crop)

                    # Confirm that the face has been saved
                    if os.path.exists(face_filename):
                        print(f"Face saved as {face_filename}")
                        face_count += 1
                        return True  # Indicate that the face was successfully saved
    return False  # Indicate that no face was detected or saved

###########################################################
# Directory where the attendance logs are saved
ATTENDANCE_LOGS_DIR = "attendance_logs/"

# Ensure the attendance logs directory exists
if not os.path.exists(ATTENDANCE_LOGS_DIR):
    os.makedirs(ATTENDANCE_LOGS_DIR)

# Global variable to hold the current log filename
current_log_filename = None

# Create a new Excel file for logging attendance
def create_new_attendance_log():
    # Create a workbook and set up sheet
    timestamp = datetime.now().strftime("%Y-%m-%d")
    filename = f"attendance_log_{timestamp}.xlsx"
    file_path = os.path.join(ATTENDANCE_LOGS_DIR, filename)  # Save inside the 'attendance_logs' directory
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Log"
    
    # Set headers
    ws["A1"] = "Student Name"
    ws["B1"] = "Timestamp"
    
    # Save the workbook
    wb.save(file_path)
    return file_path  # Return the full path to the log file

# Log attendance in the Excel sheet
def log_attendance_in_excel(student_name, filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    # Find next available row
    row = len(ws['A']) + 1  # Find the first empty row in column A
    
    # Log the student's name and timestamp in the new row
    ws[f"A{row}"] = student_name
    ws[f"B{row}"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Save the file with the updated data
    wb.save(filename)

# Load registered faces
def load_registered_faces():
    registered_faces = []
    student_names = []

    for filename in os.listdir("registered_faces"):
        if filename.endswith(".jpg") or filename.endswith(".png"):
            image = cv2.imread(os.path.join("registered_faces", filename))
            registered_faces.append(image)
            student_names.append(filename.split('_')[0])  # Extract name from filename

    return registered_faces, student_names

# Attendance frame processing
@app.route('/process_attendance_frame', methods=['POST'])
def process_attendance_frame():
    data = request.json
    frame_data = data['frame']

    # Decode Base64 image
    _, encoded_image = frame_data.split(',', 1)
    decoded_image = base64.b64decode(encoded_image)
    np_image = np.frombuffer(decoded_image, np.uint8)
    frame = cv2.imdecode(np_image, cv2.IMREAD_COLOR)

    # Load registered faces
    registered_faces, student_names = load_registered_faces()

    # Process frame with MediaPipe Face Detection
    with mp_face_detection.FaceDetection(min_detection_confidence=0.5) as face_detection:
        frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        results = face_detection.process(frame_rgb)

        if results.detections:
            for detection in results.detections:
                # Extract bounding box
                bboxC = detection.location_data.relative_bounding_box
                ih, iw, _ = frame.shape
                x, y, w, h = int(bboxC.xmin * iw), int(bboxC.ymin * ih), int(bboxC.width * iw), int(bboxC.height * ih)

                # Crop the detected face
                face_crop = frame[y:y + h, x:x + w]

                # Compare with registered faces
                for idx, registered_face in enumerate(registered_faces):
                    registered_face_resized = cv2.resize(registered_face, (face_crop.shape[1], face_crop.shape[0]))
                    difference = cv2.absdiff(face_crop, registered_face_resized)
                    similarity = np.mean(difference)

                    if similarity < 50:  # Threshold for matching
                        student_name = student_names[idx]
                        if not is_student_logged_in_current_log(student_name, current_log_filename):
                            log_attendance_in_excel(student_name, current_log_filename)
                            return jsonify({'message': f"Attendance marked for {student_name}!"})

    return jsonify({'message': "Face not recognized!"})

# Start a new attendance session and generate a new Excel file
@app.route('/start_attendance', methods=['GET'])
def start_attendance_session():
    global current_log_filename
    current_log_filename = create_new_attendance_log()
    # return jsonify({'message': 'New attendance session started.', 'filename': current_log_filename})
    return render_template('start_attendance.html')

# Check if a student has logged attendance in the current log file
def is_student_logged_in_current_log(student_name, filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    for row in range(2, len(ws['A']) + 1):  # Start checking from row 2
        if ws[f"A{row}"].value == student_name:
            return True
    return False

# Route to get the attendance log in JSON format
@app.route('/attendance_log', methods=['GET'])
def get_attendance_log():
    attendance_entries = []
    if current_log_filename:
        wb = openpyxl.load_workbook(current_log_filename)
        ws = wb.active
        for row in range(2, len(ws['A']) + 1):  # Start from row 2
            name = ws[f"A{row}"].value
            timestamp = ws[f"B{row}"].value
            attendance_entries.append({'name': name, 'timestamp': timestamp})
    return jsonify(attendance_entries)

# # Start Attendance Page
# @app.route('/start_attendance', methods=['GET'])
# def start_attendance_page():
#     return render_template('start_attendance.html')

###########################################################

@app.route('/')
def home():
    return render_template('index.html')


@app.route('/register_student', methods=['GET', 'POST'])
def register_student():
    # if request.method == 'POST':
    #     student_name = request.form['student_name']
    #     return render_template('register_student.html', student_name=student_name)
    return render_template('register_student.html')


@app.route('/process_frame', methods=['POST'])
def process_frame():
    """
    Receive video frames from the web interface and process them for face registration.
    """
    data = request.json
    student_name = data['student_name']
    frame_data = data['frame']

    # Save the image from the frame and process for face detection
    face_saved = save_image_from_base64(student_name, frame_data)

    if face_saved:
        return jsonify({'message': 'Face saved successfully!'})
    else:
        return jsonify({'message': 'No face detected. Try again.'})

##############################################################

# Endpoint to download the attendance log based on the selected date
@app.route('/download_attendance_log', methods=['GET'])
def download_attendance_log():
    # Get the date from the query parameter
    date_str = request.args.get('date')

    # Format the date into a filename
    try:
        requested_date = datetime.strptime(date_str, "%Y-%m-%d")
    except ValueError:
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD.'}), 400

    # Find the attendance log file that matches the requested date
    for filename in os.listdir(ATTENDANCE_LOGS_DIR):
        if filename.startswith(f"attendance_log_{requested_date.strftime('%Y-%m-%d')}"):
            log_file_path = os.path.join(ATTENDANCE_LOGS_DIR, filename)
            return send_file(log_file_path, as_attachment=True)

    return jsonify({'error': 'Attendance log for the selected date not found.'}), 404

@app.route('/download_attendance')
def download_attendance():
    return render_template('download.html')

if __name__ == '__main__':
    app.run(debug=True)
